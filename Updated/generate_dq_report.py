import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_report():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # STYLING DEFINITIONS
    # -------------------------------------------------------------
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold_data = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_data = Font(name="Segoe UI", size=10, color="000000")
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_section = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Status Fills and Fonts
    status_resolved = {
        "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="375623")
    }
    status_exception = {
        "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="7F6000")
    }
    status_unresolved = {
        "fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="C00000")
    }
    
    thin_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_top_bottom = Border(top=thin_side, bottom=Side(style='double', color='000000'))
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # =============================================================
    # SHEET 1: Status of Previous Issues
    # =============================================================
    ws1 = wb.active
    ws1.title = "Previous Issues Status"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A2:D2")
    ws1["A2"] = "108 EMERGENCY AMBULANCE SERVICE — PERFORMANCE MONITORING"
    ws1["A2"].font = font_title
    ws1["A2"].alignment = align_left
    
    ws1.merge_cells("A3:D3")
    ws1["A3"] = "March 2026 Data Quality Audit & Issues Resolution Report | Prepared by KPMG Advisory"
    ws1["A3"].font = font_subtitle
    ws1["A3"].alignment = align_left
    
    # Section Header
    ws1["A5"] = "I. Status of Previously Identified Issues"
    ws1["A5"].font = font_section
    ws1["A5"].alignment = align_left
    
    # Headers
    headers_ws1 = [
        "Issue ID", 
        "Previous Issue Description", 
        "Resolution Status", 
        "Current Findings & Metrics in Updated March Data"
    ]
    
    for col_idx, header in enumerate(headers_ws1, 1):
        cell = ws1.cell(row=6, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 3] else align_left
        cell.border = border_all
    
    # Data
    previous_issues_data = [
        (
            1, 
            "Case ID is Not Unique (Duplicates/Collapsed)", 
            "Resolved", 
            "0% Duplicates. All 91,315 rows in 'March 2026 Raw Data.csv' now have unique Case ID values."
        ),
        (
            2, 
            "Formula Errors (#REF!, #VALUE!, #NAME?)", 
            "Resolved", 
            "0 Errors. Cleaned completely; no instances of formula error strings were found in either file after rechecking."
        ),
        (
            3, 
            "Date & Timestamp Formats Inconsistency", 
            "Resolved (with 1 exception)", 
            "100% consistent date-time formatting in both datasets. The only exception is the invalid placeholder '00-00-0000 00:00:00' in 117 rows of raw trips."
        ),
        (
            4, 
            "Massive Data Gaps (\\N placeholders for trips)", 
            "Not Resolved", 
            "81.59% overall missing rate in raw trips. For served categories, 30.74% of Emergency Calls and 36.89% of Inter Facility Transfers are missing Vehicle Numbers. (See Sheet 2 for breakdown)."
        ),
        (
            5, 
            "Agent Connected Time Overwritten by Follow-up Calls", 
            "Not Resolved", 
            "51.48% (1,110 out of 2,156 rows) matched a follow-up call's connection time, proving the overwrite bug is still active."
        ),
        (
            6, 
            "Chronological Sequence Issues (Out-of-order stages)", 
            "Resolved", 
            "0 Anomalies. Stage transitions (connected -> assigned -> arrival) occur in correct sequence."
        ),
        (
            7, 
            "Missing District Details in Call Hits logs", 
            "Resolved for Dispatches", 
            "78.26% overall missing district rate, but this is entirely for unassigned calls. 100% of calls with an assigned vehicle have district details."
        ),
        (
            8, 
            "Blank Agent Dispositions (Call Outcomes)", 
            "Not Resolved", 
            "12.50% of calls in the Call Details report still contain the blank outcome string '---'."
        )
    ]
    
    for i, row_data in enumerate(previous_issues_data):
        row_idx = 7 + i
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_bold_data if col_idx == 1 else font_data
            cell.border = border_all
            
            # Formatting and alignment
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 3:
                cell.alignment = align_center
                # Apply custom styles based on status
                status_str = str(val).strip().upper()
                if "RESOLVED" in status_str and "NOT" not in status_str and "EXCEPTION" not in status_str:
                    cell.fill = status_resolved["fill"]
                    cell.font = status_resolved["font"]
                elif "EXCEPTION" in status_str:
                    cell.fill = status_exception["fill"]
                    cell.font = status_exception["font"]
                elif "NOT RESOLVED" in status_str:
                    cell.fill = status_unresolved["fill"]
                    cell.font = status_unresolved["font"]
                else:  # Resolved for Dispatches / other
                    cell.fill = status_resolved["fill"]
                    cell.font = status_resolved["font"]
            else:
                cell.alignment = align_left
                
            # Zebra striping on cells that are not status values
            if col_idx != 3 and row_idx % 2 == 1:
                cell.fill = fill_zebra
                
        ws1.row_dimensions[row_idx].height = 42

    # Column widths for Sheet 1
    widths_ws1 = [12, 38, 25, 78]
    for col_idx, w in enumerate(widths_ws1, 1):
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = w

    ws1.row_dimensions[2].height = 25
    ws1.row_dimensions[3].height = 18
    ws1.row_dimensions[5].height = 22
    ws1.row_dimensions[6].height = 28

    # =============================================================
    # SHEET 2: In-Depth & New Issues
    # =============================================================
    ws2 = wb.create_sheet(title="In-Depth & New Findings")
    ws2.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws2.merge_cells("A2:D2")
    ws2["A2"] = "March 2026 In-Depth Findings & Newly Identified Issues"
    ws2["A2"].font = font_title
    ws2["A2"].alignment = align_left
    
    ws2.merge_cells("A3:D3")
    ws2["A3"] = "Detailed Technical Analysis of Data Gaps, Overwrites, and Database Formatting Anomalies"
    ws2["A3"].font = font_subtitle
    ws2["A3"].alignment = align_left
    
    # -------------------------------------------------------------
    # SECTION A: Deep Dive into Gaps & Overwrites
    # -------------------------------------------------------------
    ws2["A5"] = "II. Deep Dive into Data Gaps & Overwrites (March 2026)"
    ws2["A5"].font = font_section
    ws2["A5"].alignment = align_left
    
    headers_ws2 = [
        "Issue / Ref", 
        "Metric / Core Issue Name", 
        "Affected Target File / Columns", 
        "Deep-Dive Findings & Technical/Operational Impact"
    ]
    
    for col_idx, header in enumerate(headers_ws2, 1):
        cell = ws2.cell(row=6, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx == 1 else align_left
        cell.border = border_all
        
    deep_dive_data = [
        (
            "Issue 4 (Gaps)",
            "Massive Data Gaps (Missing Vehicle Numbers)",
            "March 2026 Raw Data.csv\n- Column: Vehicle No\n- Column: Assigned_time\n- Column: scene_arrival_time",
            "• 81.59% overall missing rate in raw trips. A massive portion of this is because unserved calls are recorded in the trips log.\n• Restricting to served types, 30.74% of Emergency Calls (5,154 / 16,764) and 36.89% of Inter Facility Transfers (2,835 / 7,685) have missing vehicle numbers.\n• 90% of these gaps (4,626 Emergency, 2,548 IFT) have no remarks logged. 10% explain the gap (e.g. 'USE PVT VEHICLE', 'CANCEL BY CALLER').\n• Operational Impact: Represents either silent unlogged cancellations (should not count as 'trips') or severe data logging loss."
        ),
        (
            "Issue 5 (Overwrites)",
            "Overwritten Agent Connected Time (Follow-ups)",
            "March 2026 Raw Data.csv\n- Column: Agent CONNECTED TIME\n- Target: Caller Phone Numbers",
            "• In 51.48% of duplicate call cases (1,110 out of 2,156 analyzed), the connection time logged in raw trips was overwritten by a follow-up call.\n• Example: Caller 6200017856 called at 09:56, 09:56, and 10:10. The trip connection time was logged as 10:10. Arrival at 10:25 records response time as 15 mins (on paper) instead of 29 mins (actual), masking 14 mins of patient waiting time.\n• Operational Impact: Response times look artificially short, hiding dispatch bottlenecks."
        ),
        (
            "Issue 3 (Dates)",
            "Invalid Datetime Placeholders",
            "March 2026 Raw Data.csv\n- Column: Agent CONNECTED TIME (116 rows)\n- Column: scene_arrival_time (1 row)",
            "• Dates are 100% consistent (M/D/YYYY H:MM format) in both files, except for 117 rows containing '00-00-0000 00:00:00'.\n• Technical Impact: Standard database loaders (SQL CAST, pandas) will crash when encountering this string. Requires sanitization/conversion to NULL during ingestion."
        ),
        (
            "Issue 7 (Districts)",
            "Missing District Details in Call Hits",
            "Call Details Report March'2026.csv\n- Column: District",
            "• 78.26% overall missing district rate. However, cross-tabulation reveals that 100% of served calls (those with assigned vehicle numbers) have valid districts.\n• Technical Impact: Missing district details are restricted entirely to calls that were never dispatched."
        ),
        (
            "Issue 8 (Dispositions)",
            "Blank Agent Dispositions",
            "Call Details Report March'2026.csv\n- Column: Agent Disposition",
            "• 12.50% of call outcomes are recorded as '---'.\n• Operational Impact: Represents a lack of operational logging accountability, where 1 in 8 call outcomes remains unclassified at the agent level."
        )
    ]
    
    current_row = 7
    for row_data in deep_dive_data:
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_bold_data if col_idx == 1 else font_data
            cell.border = border_all
            cell.alignment = align_center if col_idx == 1 else align_left
            if current_row % 2 == 1:
                cell.fill = fill_zebra
        ws2.row_dimensions[current_row].height = 115 if current_row in [7, 8] else 70
        current_row += 1
        
    # -------------------------------------------------------------
    # SECTION B: New Issues
    # -------------------------------------------------------------
    current_row += 2
    ws2.cell(row=current_row, column=1, value="III. Newly Identified Data Quality Issues").font = font_section
    
    current_row += 1
    new_headers = [
        "Anomaly Code", 
        "Core Anomaly / Issue Description", 
        "Affected Target File / Columns", 
        "Technical findings & Technical/Operational Recommendations"
    ]
    
    for col_idx, header in enumerate(new_headers, 1):
        cell = ws2.cell(row=current_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx == 1 else align_left
        cell.border = border_all
    ws2.row_dimensions[current_row].height = 28
    
    new_issues_row_start = current_row + 1
    new_issues_data = [
        (
            "NEW_01 (Typo)",
            "Misspelled Column Header in Raw Trips",
            "March 2026 Raw Data.csv\n- Misspelled Column Header: 'Distict'",
            "• The district column name is misspelled as 'Distict' instead of 'District'.\n• Technical Impact: Standard database loaders, BI tools (Tableau/PowerBI), or pandas ETL scripts expecting 'District' will fail to read this column or load nulls.\n• Recommendation: Ingestion pipeline must map/rename 'Distict' to 'District'."
        ),
        (
            "NEW_02 (Precision)",
            "Loss of Precision: Scientific Notation in IDs",
            "Call Details Report March'2026.csv\n- Column: Call Reference Number\n\nMarch 2026 Raw Data.csv\n- Columns: IPD, OPD",
            "• 100% of Call Reference Numbers (299,450 rows) are saved as scientific text (e.g. 9.17E+19). Because floating-point representation only preserves 15-17 digits of precision, the last 4-5 digits of the unique 20-digit call IDs are permanently lost and filled with zeros.\n• 129 IPD entries and 12 OPD entries also suffer from scientific format truncation.\n• Technical Impact: Irreversible data corruption. Exact trip-to-call matching via Call Reference ID is impossible. Recommendation: Export database keys as explicit text/strings."
        ),
        (
            "NEW_03 (Durations)",
            "Negative Durations (Clock Synchronization Lag)",
            "Call Details Report March'2026.csv\n- Columns: Call Connect Time, Call End Time",
            "• 22 answered calls show Call End Time exactly 1 minute before Call Connect Time (e.g., Sno 5554: Connects at 13:23, ends at 13:22, duration 23s).\n• Technical Findings: Caused by terminal logging lag and truncation rounding. If a call connects at 13:22:45 and ends at 13:23:08, the database rounds Connect up to 13:23 and truncates End to 13:22.\n• Recommendation: Ignore string date differences and calculate duration directly from the raw duration seconds column."
        )
    ]
    
    for row_idx_offset, row_data in enumerate(new_issues_data):
        row_num = new_issues_row_start + row_idx_offset
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_idx, value=val)
            cell.font = font_bold_data if col_idx == 1 else font_data
            cell.border = border_all
            cell.alignment = align_center if col_idx == 1 else align_left
            if row_num % 2 == 1:
                cell.fill = fill_zebra
        ws2.row_dimensions[row_num].height = 95 if row_idx_offset == 1 else 82
        
    # Column widths for Sheet 2
    widths_ws2 = [18, 38, 35, 78]
    for col_idx, w in enumerate(widths_ws2, 1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = w
        
    ws2.row_dimensions[2].height = 25
    ws2.row_dimensions[3].height = 18
    ws2.row_dimensions[5].height = 22
    
    wb.save("/home/deveshjha/Downloads/108_KPI_Code/Updated/March_2026_Data_Quality_Report.xlsx")
    print("Report generated successfully.")

if __name__ == "__main__":
    generate_report()
