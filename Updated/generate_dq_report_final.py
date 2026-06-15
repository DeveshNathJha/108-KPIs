import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Quality Report"
    ws.views.sheetView[0].showGridLines = True
    
    # -------------------------------------------------------------
    # STYLING DEFINITIONS
    # -------------------------------------------------------------
    font_title = Font(name="Segoe UI", size=15, bold=True, color="1F4E79")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold_data = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_data = Font(name="Segoe UI", size=10, color="000000")
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    # Status Fills and Fonts
    status_resolved = {
        "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="375623")
    }
    status_exception = {
        "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="7F6000")
    }
    status_unresolved = {
        "fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "font": Font(name="Segoe UI", size=10, bold=True, color="C00000")
    }
    
    thin_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # =============================================================
    # TITLE BLOCK
    # =============================================================
    ws.merge_cells("A2:D2")
    ws["A2"] = "108 EMERGENCY AMBULANCE SERVICE — PERFORMANCE MONITORING"
    ws["A2"].font = font_title
    ws["A2"].alignment = align_left
    
    ws.merge_cells("A3:D3")
    ws["A3"] = "March 2026 Data Quality Audit & Issues Resolution Report | Prepared by KPMG Advisory"
    ws["A3"].font = font_subtitle
    ws["A3"].alignment = align_left
    
    # =============================================================
    # TABLE 1: Status of Previously Identified Issues
    # =============================================================
    ws["A5"] = "I. Status of Previously Identified Issues"
    ws["A5"].font = font_section
    ws["A5"].alignment = align_left
    
    headers_t1 = [
        "Issue ID", 
        "Previously Identified Issue Description", 
        "Resolution Status", 
        "Current Findings & Metrics in Updated March Data"
    ]
    
    for col_idx, header in enumerate(headers_t1, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 3] else align_left
        cell.border = border_all
    ws.row_dimensions[6].height = 28
    
    previous_issues_data = [
        (
            1, 
            "Case ID is Not Unique (Duplicates/Collapsed)", 
            "Resolved", 
            "0% Duplicates. All 91,315 rows in March 2026 Raw Data.csv now have unique Case ID values."
        ),
        (
            2, 
            "Formula Errors (#REF!, #VALUE!, #NAME?)", 
            "Resolved", 
            "0 Errors. Cleaned completely; no instances of formula error strings were found in either file after rechecking."
        ),
        (
            3, 
            "Date & Timestamp Formats Inconsistency", 
            "Resolved (with 1 exception)", 
            "100% consistent date-time formatting in both datasets. The only exception is the invalid placeholder 00-00-0000 00:00:00 in 117 rows of raw trips."
        ),
        (
            4, 
            "Massive Data Gaps (\\N placeholders for trips)", 
            "Not Resolved", 
            "81.59% overall missing rate in raw trips. For served categories, 30.74% of Emergency Calls and 36.89% of Inter Facility Transfers are missing Vehicle Numbers. (See detailed breakdown below)."
        ),
        (
            5, 
            "Agent Connected Time Superseded by Subsequent Calls", 
            "Not Resolved", 
            "51.48% (1,110 out of 2,156 rows) matched a subsequent call's connection time, proving the connection time replacement behavior is still active."
        ),
        (
            6, 
            "Chronological Sequence Issues (Out-of-order stages)", 
            "Resolved", 
            "0 Anomalies. Stage transitions (connected -> assigned -> arrival) occur in correct sequence."
        ),
        (
            7, 
            "Missing District Details in Call Hits logs", 
            "Resolved for Dispatches", 
            "78.26% overall missing district rate, but this is entirely for unassigned calls. 100% of calls with an assigned vehicle have district details."
        ),
        (
            8, 
            "Blank Agent Dispositions (Call Outcomes)", 
            "Not Resolved", 
            "12.50% of calls in the Call Details report still contain the blank outcome string '---'."
        )
    ]
    
    for i, row_data in enumerate(previous_issues_data):
        row_idx = 7 + i
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_bold_data if col_idx == 1 else font_data
            cell.border = border_all
            
            # Formatting and alignment
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 3:
                cell.alignment = align_center
                status_str = str(val).strip().upper()
                if "RESOLVED" in status_str and "NOT" not in status_str and "EXCEPTION" not in status_str:
                    cell.fill = status_resolved["fill"]
                    cell.font = status_resolved["font"]
                elif "EXCEPTION" in status_str:
                    cell.fill = status_exception["fill"]
                    cell.font = status_exception["font"]
                elif "NOT RESOLVED" in status_str:
                    cell.fill = status_unresolved["fill"]
                    cell.font = status_unresolved["font"]
                else:  # Resolved for Dispatches / other
                    cell.fill = status_resolved["fill"]
                    cell.font = status_resolved["font"]
            else:
                cell.alignment = align_left
                
            # Zebra striping on cells that are not status values
            if col_idx != 3 and row_idx % 2 == 1:
                cell.fill = fill_zebra
                
        ws.row_dimensions[row_idx].height = 42
        
    # =============================================================
    # TABLE 2: Newly Identified Issues & Critical Technical Gaps
    # =============================================================
    start_t2_section = 17
    ws.cell(row=start_t2_section, column=1, value="II. Newly Identified Issues & Technical Gaps").font = font_section
    ws.cell(row=start_t2_section, column=1).alignment = align_left
    
    headers_t2 = [
        "Code / Ref", 
        "Newly Identified Issue & Critical Gap Description", 
        "Affected Target File & Columns", 
        "Detailed Findings, Technical Impact & Professional Recommendations"
    ]
    
    header_t2_row = start_t2_section + 1
    for col_idx, header in enumerate(headers_t2, 1):
        cell = ws.cell(row=header_t2_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 3] else align_left
        cell.border = border_all
    ws.row_dimensions[header_t2_row].height = 28
    
    new_issues_data = [
        (
            "NEW_01 (Typo)",
            "Misspelled Column Header in Raw Trips",
            "March 2026 Raw Data.csv\n- Misspelled Column Header: 'Distict'",
            "• Column name is misspelled as 'Distict' instead of 'District'.\n• Technical Impact: Standard database loaders, BI tools, or pandas scripts expecting 'District' will fail to read this column.\n• Recommendation: Ingestion pipeline must implement a rename mapping."
        ),
        (
            "NEW_02 (Precision)",
            "Loss of Precision: Scientific Notation in IDs",
            "Call Details Report March'2026.csv\n- Column: Call Reference Number\n\nMarch 2026 Raw Data.csv\n- Columns: IPD, OPD",
            "• 100% of Call Reference Numbers (299,450 rows) are saved as scientific text (e.g. 9.17E+19). IPD/OPD columns also contain truncated floating-point values.\n• Technical Impact: Floating-point conversion permanently deletes the last 4-5 digits of these 20-digit unique IDs, causing collisions and making individual transaction-level tracking impossible.\n• Recommendation: Export database keys as explicit text/strings rather than floats."
        ),
        (
            "NEW_03 (Durations)",
            "Negative Durations (Clock Synchronization Lag)",
            "Call Details Report March'2026.csv\n- Columns: Call Connect Time, Call End Time",
            "• 22 answered calls show Call End Time exactly 1 minute before Call Connect Time (e.g., Sno 5554: Connects at 13:23, ends at 13:22, duration 23s).\n• Technical Findings: Caused by terminal logging lag and truncation rounding. If a call connects at 13:22:45 and ends at 13:23:08, the database rounds Connect up to 13:23 and truncates End to 13:22.\n• Recommendation: Re-calculate durations from raw call center duration seconds column."
        ),
        (
            "GAP_01 (Trips)",
            "Critical Data Gaps in Serviced Categories",
            "March 2026 Raw Data.csv\n- Columns: Vehicle No, Assigned_time, scene_arrival_time",
            "• 30.74% of Emergency Calls (5,154 rows) and 36.89% of IFTs (2,835 rows) are missing vehicle numbers.\n• Technical Impact: 90% of these gaps lack operators' remarks, suggesting either unlogged service delivery or silent operational cancellations.\n• Recommendation: Force operator remarks validation when vehicle dispatch fields are left empty."
        )
    ]
    
    t2_data_start = header_t2_row + 1
    for offset, row_data in enumerate(new_issues_data):
        row_num = t2_data_start + offset
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = font_bold_data if col_idx == 1 else font_data
            cell.border = border_all
            cell.alignment = align_center if col_idx in [1, 3] else align_left
            if row_num % 2 == 1:
                cell.fill = fill_zebra
        ws.row_dimensions[row_num].height = 95
        
    # Set column widths
    widths = [14, 38, 25, 78]
    for col_idx, w in enumerate(widths, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = w
        
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[5].height = 22
    
    wb.save("/home/deveshjha/Downloads/108_KPI_Code/Updated/March_2026_Data_Quality_Report_Final.xlsx")
    print("Final Report generated successfully on a single sheet.")

if __name__ == "__main__":
    generate_report()
